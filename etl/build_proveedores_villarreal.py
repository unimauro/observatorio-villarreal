import openpyxl, json, os
HERE=os.path.dirname(os.path.abspath(__file__))
CONOSCE=os.environ.get('CONOSCE_DIR','/Users/unimauro/Documents/Repos/observatorio-sanmarcos/etl/conosce')
CODIGO='001902'; RUC_ENT='20170934289'  # UNFV
def tp(ruc):
    p=str(ruc)[:2]
    return 'natural' if p in ('10','15') else 'empresa'

# dueno/representante (via datosperu.org). null = sin fuente confiable. (rellenar luego)
duenos={}

rows=[]
for Y in (2023,2024,2025):
    wb=openpyxl.load_workbook(f"{CONOSCE}/CONOSCE_ADJUDICACIONES{Y}_0.xlsx", read_only=True)
    ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); next(it)
    for r in it:
        if str(r[0])==CODIGO and str(r[1])==RUC_ENT:
            rows.append(r)
    wb.close()
print("item rows:", len(rows))

agg={}
for r in rows:
    ruc=str(r[19]); monto=r[15] or 0
    a=agg.setdefault(ruc,{'nombre':r[20],'ruc':ruc,'monto':0.0,'convs':set(),
        'tipos':{},'objeto':{},'tipo_persona':tp(ruc),'tipo_proveedor_seace':r[21]})
    a['monto']+=float(monto)
    a['convs'].add(r[5])
    a['tipos'][r[7]]=a['tipos'].get(r[7],0)+1
    a['objeto'][r[6]]=a['objeto'].get(r[6],0)+1
provs=[]
for ruc,a in agg.items():
    d,f=duenos.get(ruc,(None,None))
    provs.append({'nombre':a['nombre'],'ruc':ruc,'monto':round(a['monto'],2),
        'n':len(a['convs']),'tipos':a['tipos'],'objeto':a['objeto'],
        'tipo_persona':a['tipo_persona'],'tipo_proveedor_seace':a['tipo_proveedor_seace'],
        'dueno':d,'fuente_dueno':f})
provs.sort(key=lambda x:-x['monto'])
monto_total=round(sum(p['monto'] for p in provs),2)
emp=[p for p in provs if p['tipo_persona']=='empresa']
nat=[p for p in provs if p['tipo_persona']=='natural']
all_convs=set(r[5] for r in rows)
top_personas=sorted([{'nombre':p['nombre'],'ruc':p['ruc'],'monto':p['monto'],'n':p['n']} for p in nat],key=lambda x:-x['monto'])
out={
 '_meta':{
  'fuente':'OECE/OSCE - CONOSCE Datos Abiertos, reporte de Adjudicaciones (buena pro por item)',
  'fuente_url':'https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/adjudicaciones/',
  'entidad':'Universidad Nacional Federico Villarreal (UNFV) - pliego 524',
  'ruc':RUC_ENT,
  'codigoentidad_conosce':int(CODIGO),
  'periodo':'2023-2025',
  'extraido':'2026-07',
  'unidad_monto':'Soles (PEN), monto adjudicado por item',
  'nota':"Agregado desde reportes anuales CONOSCE de Adjudicaciones (nivel item de buena pro), archivos CONOSCE_ADJUDICACIONES{2023,2024,2025}_0.xlsx, filtrando codigoentidad 001902 / RUC 20170934289 (UNIVERSIDAD NACIONAL FEDERICO VILLARREAL - UNFV, entidad de Gobierno Nacional). 'monto' = suma de monto_adjudicado_item_soles. 'n' = numero de procesos (codigoconvocatoria) distintos en que ese proveedor obtuvo buena pro. 'tipos' = procesos por tipo de proceso de seleccion. 'objeto' = items por objeto contractual (Bien/Servicio/Obra). tipo_persona por prefijo de RUC (20=empresa, 10/15=natural); consorcios y no domiciliados clasificados como empresa. NO incluye ordenes de compra <8 UIT (dataset aparte)."
 },
 'totales':{
  'monto_total':monto_total,
  'n_proveedores':len(provs),
  'n_procesos':len(all_convs),
  'n_empresas':len(emp),
  'n_personas_naturales':len(nat),
  'monto_empresas':round(sum(p['monto'] for p in emp),2),
  'monto_personas_naturales':round(sum(p['monto'] for p in nat),2),
 },
 'top_personas':top_personas,
 'proveedores':provs,
}
json.dump(out, open(os.path.join(HERE,'..','data','proveedores-villarreal.json'),'w'),
          ensure_ascii=False, separators=(",",":"))
print("proveedores:", len(provs), "monto_total:", monto_total, "procesos:", len(all_convs))
print("empresas:", len(emp), "naturales:", len(nat))
print("--- TOP 15 ---")
for p in provs[:15]:
    print(f"  {p['monto']:>14,.2f}  {p['ruc']}  {p['tipo_persona']:8}  {p['nombre']}")
